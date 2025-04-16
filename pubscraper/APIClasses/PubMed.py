import requests
import json
import time
import logging
from dateutil.parser import parse
from ratelimit import limits, sleep_and_retry
import os

from pubscraper.APIClasses.Base import Base
import pubscraper.config as config

logger = logging.getLogger(__name__)

class PubMed(Base):
    def __init__(self):
        self.search_url = config.PUBMED_SEARCH_URL
        self.fetch_url = config.PUBMED_FETCH_URL
        logging.debug(f"PubMed API rate limit: 2 requests per second (API limit is 3/second)")

    @sleep_and_retry
    @limits(calls=2, period=1)  # Use 2 requests per second for safety (PubMed API limit is 3 per second)
    def _make_request(self, url, params):
        try:
            response = requests.get(url, params=params, timeout=10)
        except requests.exceptions.RequestException as e:
            logging.error(f"Error fetching data from {url}: {e}")
            raise
        response.raise_for_status()
        return response

    def _get_UIDs_by_author(self, author_name, rows=10):
        """
        Retrieve a given author's UID publications
        :param author_name: name of author in format "Last First [Middle]"
        :param rows: number of results to return (default is 10)
        :return: A list of UIDs corresponding to papers written by the author
        """
        if not author_name or author_name == "":
            logging.debug("Skipping empty author name")
            return None

        # Split the name and format for PubMed's search
        split_name = author_name.split()
        if len(split_name) < 2:
            logging.warning(f"Invalid author name format: {author_name}")
            return None
        
        # Try different search formats
        # Format 1: Last+First+Middle[Full Author Name]
        name_search1 = f"{split_name[0]}+{'+'.join(split_name[1:])}[Full Author Name]"
        # Format 2: Last+First[Author]
        name_search2 = f"{split_name[0]}+{split_name[1]}[Author]"
        
        logging.debug(f"Trying search formats:")
        logging.debug(f"1: {name_search1}")
        logging.debug(f"2: {name_search2}")

        for search_term in [name_search1, name_search2]:
            params = {
                "db": "pubmed",  # Try pubmed instead of pmc
                "term": search_term,
                "retmax": rows,
                "retmode": "JSON",
            }

            try:
                # Log the full URL being called
                query_url = f"{self.search_url}?{'&'.join(f'{k}={v}' for k, v in params.items())}"
                logging.debug(f"Making request to: {query_url}")
                
                response = self._make_request(self.search_url, params=params)
                data = response.json()
                
                # Log the response
                logging.debug(f"Response: {json.dumps(data, indent=2)}")
                
                if "esearchresult" in data and "idlist" in data["esearchresult"]:
                    id_list = data["esearchresult"]["idlist"]
                    if id_list:
                        logging.info(f"Found {len(id_list)} publications for {author_name}")
                        return id_list
                
            except Exception as e:
                logging.error(f"PubMed API Request error: {e}")
                continue

        logging.info(f"No publications found for author: {author_name}")
        return None

    def _get_publication_details(self, UIDs, author_name=None):
        """
        Get detailed publication information using efetch
        :params UIDs: list of UIDs
        :params author_name: name of author to check affiliations for
        :return: list of publication dictionaries
        """
        if not UIDs:
            return None

        params = {
            "db": "pubmed",
            "id": ",".join(UIDs),
            "retmode": "xml"
        }

        try:
            response = self._make_request(self.fetch_url, params=params)
            from xml.etree import ElementTree as ET
            root = ET.fromstring(response.text)
            
            publications = []
            for article in root.findall(".//PubmedArticle"):
                try:
                    # Extract only necessary information
                    title = article.find(".//ArticleTitle").text
                    journal = article.find(".//Journal/Title").text
                    doi = next((id_elem.text for id_elem in article.findall(".//ArticleId") if id_elem.get("IdType") == "doi"), "")
                    
                    # Get publication date
                    pub_date = article.find(".//PubDate")
                    year = pub_date.find("Year")
                    month = pub_date.find("Month")
                    day = pub_date.find("Day")
                    publication_date = f"{year.text if year is not None else ''}-{month.text if month is not None else ''}-{day.text if day is not None else ''}"
                    
                    # Get authors and affiliations
                    authors = []
                    affiliations = []
                    for author in article.findall(".//Author"):
                        last_name = author.find("LastName")
                        fore_name = author.find("ForeName")
                        author_name = " ".join(filter(None, [fore_name.text if fore_name is not None else '', last_name.text if last_name is not None else '']))
                        authors.append(author_name)

                        # Get affiliations for this author
                        aff_list = author.findall(".//AffiliationInfo/Affiliation")
                        author_affiliations = [aff.text for aff in aff_list if aff.text]
                        
                        if author_affiliations:
                            affiliations.append({
                                "author": author_name,
                                "affiliations": author_affiliations
                            })

                    # Check for UT system affiliation
                    if not self._check_ut_affiliation(affiliations, author_name):
                        continue

                    pub = {
                        "from": "PubMed",
                        "journal": journal,
                        "publication_date": publication_date,
                        "title": title,
                        "authors": ",".join(authors),
                        "affiliations": affiliations,
                        "doi": doi
                    }
                    publications.append(pub)
                    
                except Exception as e:
                    logging.warning(f"Error processing article: {e}")
                    continue

            if publications:
                logging.info(f"Successfully processed {len(publications)} publications")
                return publications
            else:
                logging.warning("No publications with UT system affiliations found")
                return None

        except Exception as e:
            logging.error(f"Error fetching data from PubMed: {e}")
            return None

    def _check_ut_affiliation(self, affiliations, author_name=None):
        """
        Check if any affiliation contains UT system keywords
        If author_name is provided, check only that specific author's affiliation
        :param affiliations: List of affiliation dictionaries
        :param author_name: Optional name of the specific author to check
        :return: Boolean indicating if UT system affiliation is present
        """
        if not affiliations:
            return False
        
        for author_info in affiliations:
            # If author_name is provided, only check that specific author
            if author_name and not author_info['author'].lower().startswith(author_name.lower().split()[0].lower()):
                continue
                
            for affiliation in author_info.get('affiliations', []):
                if not affiliation:
                    continue
                if 'university of texas' in affiliation.lower():
                    logging.debug(f"Found UT affiliation for {author_info['author']}: {affiliation}")
                    return True
        
        return False

    def get_publications_by_author(self, author_name, rows=10):
        """
        Given the name of an author, search PubMed for works written by that author
        :params author_name: name of author to search
        :params rows: maximum number of publications to return (default is 10)
        :return: a list of publication objects/dicts holding data for each publication
        """
        logging.debug(f"Fetching publications for author: {author_name}")
        UIDs = self._get_UIDs_by_author(author_name, rows)
        if not UIDs:
            logging.info(f"No publications found for {author_name}")
            return None
        
        publications = self._get_publication_details(UIDs, author_name)
        if publications:
            logging.debug(f"Successfully retrieved {len(publications)} publications")
        return publications


def search_multiple_authors(authors, rows=10):
    """
    Search PubMed Central for works written by multiple authors
    :params authors: list of author names
    :params rows: maximum number of publications to return per author (default is 10)
    :return: a dict {author_name: {summary_info}} for each author
    """
    pubmed = PubMed()
    all_results = {}

    for author in authors:
        logging.debug(f"Searching for publications by {author}...")
        if not author or author == "None None":  # Skip empty or None None authors
            logging.warning(f"Skipping invalid author name: {author}")
            continue
        try:
            publications = pubmed.get_publications_by_author(author, rows)
            if publications:  # Only add authors with valid publications
                all_results[author] = publications
        except Exception as e:
            logging.error(f"Error fetching data for {author}: {e}")
        time.sleep(0.4)  # avoids RESPONSE 429 (rate limit violation)

    return all_results


def main():
    author_names = input("Enter author names (comma-separated): ").split(",")
    author_names = [name.strip() for name in author_names]

    # Add debug information
    try:
        import pubscraper.config as config
        print(f"Current WS_NAME in config: {config.WS_NAME}")
        
        # If you're using openpyxl, add this:
        from openpyxl import load_workbook
        wb = load_workbook('example_UTRC_report_October-2024.xlsx')
        print(f"Available worksheets: {wb.sheetnames}")
        
        results = search_multiple_authors(author_names)
        print(json.dumps(results, indent=2))
    except Exception as e:
        print(f"Error occurred: {type(e).__name__}: {str(e)}")
        print(f"Current working directory: {os.getcwd()}")


if __name__ == "__main__":
    main()
