import json
import logging
import time
import os
import tablib

from openpyxl import load_workbook
from dateutil.parser import parse

import click
from click_loglevel import LogLevel

from pubscraper.version import __version__
import pubscraper.config as config

from pubscraper.APIClasses.PubMed import PubMed
from pubscraper.APIClasses.arXiv import ArXiv
from pubscraper.APIClasses.MDPI import MDPI
from pubscraper.APIClasses.Elsevier import Elsevier
from pubscraper.APIClasses.Springer import Springer
from pubscraper.APIClasses.Wiley import Wiley
from pubscraper.APIClasses.CrossRef import CrossRef
from pubscraper.APIClasses.PLOS import PLOS


LOG_FORMAT = config.LOGGER_FORMAT_STRING
LOG_LEVEL = config.LOGGER_LEVEL
logging.basicConfig(level=LOG_LEVEL, format=LOG_FORMAT)
logger = logging.getLogger(__name__)

APIS = {
    "PubMed": PubMed(),
    "ArXiv": ArXiv(),
    "MDPI": MDPI(),
    "Elsevier": Elsevier(),
    "Springer": Springer(),
    "Wiley": Wiley(),
    "CrossRef": CrossRef(),
    "PLOS": PLOS(),
}


def set_logging_level(ctx, param, value):
    """
    Callback function for click that sets the logging level
    """
    logger.setLevel(value)
    return value


def set_log_file(ctx, param, value):
    """
    Callback function for click that sets a log file
    """
    if value:
        fileHandler = logging.FileHandler(value, mode="w")
        logFormatter = logging.Formatter(LOG_FORMAT)
        fileHandler.setFormatter(logFormatter)
        logger.addHandler(fileHandler)
    return value


def list_configured_apis(ctx, param, value):
    """
    Callback function for click that lists available APIs
    """
    if value:
        click.secho("Available endpoints:", underline=True)
        for endpoint in APIS.keys():
            click.secho(f"  {endpoint}", fg="blue")
        ctx.exit()


@click.command()
@click.version_option(__version__)
@click.option(
    "--log-level",
    type=LogLevel(),
    default=logging.INFO,
    is_eager=True,
    callback=set_logging_level,
    help="Set the log level",
    show_default=True,
)
@click.option(
    "--log-file",
    type=click.Path(writable=True),
    is_eager=True,
    callback=set_log_file,
    help="Set the log file",
)
@click.option(
    "-i",
    "--input_file",
    type=click.Path(exists=True),
    default="example_input.xlsx",
    help="Specify input file",
)
@click.option("-o", "--output_file", default="output", help="Specify output file")
@click.option(
    "-n",
    "--number",
    type=int,
    default=10,
    help="Specify max number of publications to receive for each author",
)
@click.option(
    "--apis",
    "-a",
    type=click.Choice(
        [api for api in APIS.keys()],
        case_sensitive=False,
    ),
    multiple=True,
    default=[api for api in APIS.keys()],
    show_default=True,
    help="Specify APIs to query",
)
# TODO: I don't like the help message saying 'available' for querying, rephrase for clarity
@click.option(
    "--list",
    "list_apis",
    is_flag=True,
    default=False,
    is_eager=True,
    callback=list_configured_apis,
    help="Display APIs configured for search queries",
)
@click.option(
    "--format",
    "-f",
    type=click.Choice(
        ["json", "csv", "xlsx"],
        case_sensitive=False,
    ),
    default="json",
    show_default=True,
    help="Select the output format from: csv, xlsx, or json.",
)
@click.option(
    "--cutoff_date",
    "-cd",
    type=str,
    default=None,
    show_default=True,
    help="Specify the latest date to pull publications. Example input: 2024 or 2024-05 or 2024-05-10.",
)

# TODO: batch author names to circumvent rate limits?
def main(
    log_level,
    log_file,
    input_file,
    number,
    output_file,
    apis,
    list_apis,
    format,
    cutoff_date,
):
    logger.debug(f"Logging is set to level {logging.getLevelName(log_level)}")
    if log_file:
        logger.debug(f"Writing logs to {log_file}")

    logger.info(f"Querying the following APIs:\n{(", ").join(apis)}")
    try:
        authors_workbook = load_workbook(filename=input_file, read_only=True)
        worksheet = authors_workbook["Sheet1"]
        rows = worksheet.rows

        name_dict = {}
        if worksheet.max_row > 1:
            next(rows)  # skip header row
            for row in rows:
                institution = row[0].value
                author_name = f"{row[1].value} {row[2].value}"
                name_dict[author_name] = [author_name, institution]

        logging.debug(f"number of names in name_dict: {len(name_dict.keys())}")
    except FileNotFoundError:
        logger.error(f"Couldn't read input file {input_file}, exiting")
        exit(1)

    logger.debug(f"Querying the following APIs: {apis}")
    logger.debug(f"Requesting {number} publications for each author")

    authors_and_pubs = []

    for author in name_dict.keys():
        results = {author: []}
        # FIXME: we should filter by date before the API queries (if the API supports date filtering)
        authors_pubs = []
        for api_name in apis:
            api = APIS[api_name]
            pubs_found = api.get_publications_by_author(author, number)
            if pubs_found:
                for pub in pubs_found:
                    publication_date_str = pub.get("publication_date", "")

                    # If a cutoff date is provided, check if the publication date is after it
                    if cutoff_date:
                        publication_date = (
                            parse(publication_date_str).strftime("%Y-%m-%d")
                            if publication_date_str
                            else ""
                        )
                        if publication_date and publication_date > cutoff_date:
                            authors_pubs.append(pub)
                    else:
                        authors_pubs.append(pub)

        results.update({author: authors_pubs})
        authors_and_pubs.append(results)
        time.sleep(0.4)

    """
    Using TabLib to format data in specified format
    """
    logger.debug(f"Results: {(json.dumps(authors_and_pubs, indent=2))}")
    logger.info(f"Exporting the dataset in the specified format: {format} ")

    try:
        os.remove(output_file)
        logger.debug(f"successfully removed {output_file}")
    except Exception:
        logger.warning(f"could not remove {output_file}")

    dataset = tablib.Dataset()

    dataset.headers = [
        "From",
        "Author",
        "DOI",
        "Journal",
        "Content Type",
        "Publication Date",
        "Title",
        "Authors",
    ]

    # Loop through each author and their publications in authors_and_pubs
    for author_result in authors_and_pubs:
        for (
            author,
            publications,
        ) in author_result.items():  # Use .items() to unpack dictionary
            for pub in publications:
                if isinstance(pub, dict):  # Only process dictionary entries
                    # Safely fetch values using .get to avoid KeyError, defaulting to 'N/A' if the key is missing
                    dataset.append(
                        [
                            pub.get("from", "N/A"),
                            author,
                            pub.get("doi", "N/A"),
                            pub.get("journal", "N/A"),
                            pub.get("content_type", "N/A"),
                            pub.get("publication_date", "N/A"),
                            pub.get("title", "N/A"),
                            pub.get("authors", "N/A"),
                        ]
                    )

    if format == "xlsx":
        with open(f"output.{format}", "wb") as f:
            f.write(dataset.export("xlsx"))
    else:
        with open(f"output.{format}", "w") as f:
            if format == "csv":
                f.write(dataset.export("csv"))
            elif format == "json":
                json.dump(authors_and_pubs, f, indent=4)

    logger.info(f"Data successfully exported to {output_file}.{format}")

    return 0


if __name__ == "__main__":
    main()
